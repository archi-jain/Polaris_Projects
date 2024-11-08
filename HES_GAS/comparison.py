import os
import pandas as pd
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Load the two CSV files using a file dialog
def load_csv():
    root = tk.Tk()
    root.withdraw()  # Hide the main Tkinter window

    # Ask the user to select the first CSV file
    file1 = filedialog.askopenfilename(title="Select the first CSV file", filetypes=[("CSV files", "*.csv")])
    if not file1:
        print("No file selected. Exiting.")
        return None, None

    # Ask the user to select the second CSV file
    file2 = filedialog.askopenfilename(title="Select the second CSV file", filetypes=[("CSV files", "*.csv")])
    if not file2:
        print("No file selected. Exiting.")
        return None, None

    # Load the CSVs into dataframes
    df1 = pd.read_csv(file1)
    df2 = pd.read_csv(file2)

    return df1, df2

# Combine columns side by side and compare values for differences
def compare_and_merge(df1, df2, key_column, exclude_columns):
    merged_data = pd.DataFrame()
    highlight_data = pd.DataFrame()

    for column in df1.columns:
        if column in exclude_columns:
            # For excluded columns, write the column only once
            merged_data[column] = df1[column]
        else:
            # For other columns, merge the columns from both files side by side
            merged_data[column + '_file1'] = df1[column]
            merged_data[column + '_file2'] = df2[column]
            # Store the comparison (True/False) in a separate DataFrame for highlighting
            highlight_data[column] = df1[column] != df2[column]

    return merged_data, highlight_data

# Save the DataFrame to Excel and highlight the differences
def save_and_highlight_excel(merged_data, highlight_data, meter_number):
     # Determine the Downloads folder path
    downloads_folder = os.path.join(os.path.expanduser("~"), "Downloads")

    # Create an Excel writer object
    excel_file = os.path.join(downloads_folder, f'comparison_{meter_number}.xlsx')
    merged_data.to_excel(excel_file, index=False, engine='openpyxl')

    # Load the workbook and select the active sheet
    wb = load_workbook(excel_file)
    ws = wb.active

    # Define the yellow fill color for highlighting differences
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Iterate through the DataFrame and apply yellow fill where differences exist
    for row in range(2, len(merged_data) + 2):  # Skip header row
        for col in range(1, len(merged_data.columns) + 1):
            # Get the original column name (removing _file1/_file2 suffix for matching)
            header = ws.cell(row=1, column=col).value
            if header.endswith('_file1') or header.endswith('_file2'):
                original_column = header.rsplit('_', 1)[0]
                # Apply yellow fill if there's a difference in the respective highlight column
                if original_column in highlight_data.columns and highlight_data.loc[row-2, original_column]:
                    ws.cell(row=row, column=col).fill = yellow_fill

    # Save the workbook
    wb.save(excel_file)

if __name__ == "__main__":
    # Load CSV files through file dialog
    df1, df2 = load_csv()

    if df1 is not None and df2 is not None:
        # Define columns to exclude
        exclude_columns = ['meter_number', 'PF_Serial_Number', 'BP_Number', 'GA', 'project', 'battery_level_percentage']  # Replace with actual columns to exclude

        # Define the key column for merging and filename purposes
        key_column = 'meter_number'  # Replace with the actual key column name if different

        # Ensure that meter_number is consistent between both dataframes
        if df1[key_column].equals(df2[key_column]):
            # Get the meter number for the filename
            meter_number = str(df1[key_column].iloc[0])  # Assuming all rows have the same meter number

            # Combine columns side by side and compare values
            merged_data, highlight_data = compare_and_merge(df1, df2, key_column, exclude_columns)

            # Save the output to Excel with differences highlighted in yellow
            save_and_highlight_excel(merged_data, highlight_data, meter_number)

            print(f"Comparison completed. Differences saved and highlighted in 'comparison_result_{meter_number}.xlsx'.")
        else:
            print("Meter numbers do not match between the two files.")
    else:
        print("Operation cancelled or files not loaded correctly.")
